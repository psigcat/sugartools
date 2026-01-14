from qgis.PyQt.QtWidgets import QComboBox
from qgis.core import Qgis, QgsProject, QgsPointXY, QgsDistanceArea, QgsVectorLayer, QgsFeature, QgsGeometry, QgsLayerTreeLayer, QgsFields, QgsField
from qgis.PyQt.QtCore import QVariant

import os
import math
import csv
import xlrd
import openpyxl


from .utils import utils


SYMBOLOGY_DIR = "qml"
COMBO_SELECT = "(Select)"
FIELDS_DEFAULT = {
    "refitting_part": "num_pieza",
    "refitting_coordx": "coord_x",
    "refitting_coordy": "coord_y",
    "refitting_coordz": "coord_z",
    "refitting_origin": "Origen",
    "refitting_target": "Destino",
    "refitting_class": "Clase Rem",
    "refitting_labels": "num_pieza",
    "refitting_colors": "color_tremon"
}
FIELDS_MANDATORY_REMOUNTING = ["refitting_excel", "refitting_sheet", "refitting_part", "refitting_coordx", "refitting_coordy", "refitting_coordz", "refitting_origin", "refitting_target", "refitting_class", "refitting_labels"]
FIELDS_MANDATORY_REMOUNTING_CSV = ["refitting_excel", "refitting_part", "refitting_coordx", "refitting_coordy", "refitting_coordz", "refitting_origin", "refitting_target", "refitting_class", "refitting_labels"]
CSV_PARAMS = '?maxFields=20000&detectTypes=yes&crs=EPSG:25831&spatialIndex=no&subsetIndex=no&watchFile=no'


class RefittingTool():
    def __init__(self, parent):
        """Constructor."""

        self.parent = parent
        self.extension = None
        self.book = None
        self.sheet = None
        self.sheet_name = None
        self.sheetrows = 0
        self.table = []
        self.tablekeys = []
        self.col_indexes = {}
        self.parts = []
        self.points = {}
        self.lines = []
        self.lines_attr = []

        self.utils = utils(self.parent)


    def setup(self):
        """ load initial parameters """

        self.parent.dlg.refitting_excel.fileChanged.connect(self.load_excel)


    def load_excel(self):
        """ load sheets from excel """

        if self.book:
            if self.extension == "xls":
                self.parent.dlg.refitting_sheet.currentIndexChanged.disconnect(self.load_xls_sheet)
            elif self.extension == "xlsx":
                self.parent.dlg.refitting_sheet.currentIndexChanged.disconnect(self.load_xlsx_sheet)
            elif self.extension == "csv":
                self.parent.dlg.refitting_sheet.currentIndexChanged.disconnect(self.load_csv_sheet)

        self.parent.dlg.refitting_sheet.clear()
        self.parent.dlg.refitting_sheet.addItem(COMBO_SELECT)
        self.clear_parts()

        file_path = self.parent.dlg.refitting_excel.filePath()
        self.extension = os.path.basename(file_path).split(".")[1]

        if self.extension == "xls":
            self.load_xls(file_path)
        elif self.extension == "xlsx":
            self.load_xlsx(file_path)
        elif self.extension == "csv":
            self.load_csv(file_path)
        else:
            self.parent.dlg.messageBar.pushMessage(f"Only .xls, .xlsx and .csv files are supported, {self.extension} selected.", level=Qgis.Warning)
            return


    def load_xls(self, file_path):
        """ load sheets from XLS """

        self.book = xlrd.open_workbook(file_path)
        print("The number of worksheets is {0}".format(self.book.nsheets))
        print("Worksheet name(s): {0}".format(self.book.sheet_names()))

        for index in range(self.book.nsheets):
            sheet = self.book.sheet_by_index(index)
            self.parent.dlg.refitting_sheet.addItem(sheet.name)

        self.parent.dlg.refitting_sheet.currentIndexChanged.connect(self.load_xls_sheet)


    def load_xls_sheet(self):
        """ load data from excel sheet """

        index = self.parent.dlg.refitting_sheet.currentIndex() - 1
        self.sheet = self.book.sheet_by_index(index)
        self.sheet_name = self.sheet.name
        self.sheetrows = self.sheet.nrows
        print("Sheet '{0}' has {1} rows and {2} columns".format(self.sheet_name, self.sheet.nrows, self.sheet.ncols))

        # load column indexes
        for column_index in range(self.sheet.ncols):
            name = self.sheet.cell_value(0, column_index)
            self.col_indexes[name] = column_index
            self.fill_parts(name)

        self.select_default()

        # load all data into dict in order to save it later to csv file
        self.tablekeys = list(self.col_indexes.keys())
        for rx in range(1, self.sheet.nrows):
            row = {}
            for ry in range(self.sheet.ncols):
                col = self.tablekeys[ry]
                val = self.sheet.cell_value(rx, ry)
                type = self.sheet.cell_type(rx, ry)
                if type == 2: #number
                    val = int(val)
                row[col] = val
            self.table.append(row)
        #print(self.table)


    def load_xlsx(self, file_path):
        """ load sheets from XLS """

        self.book = openpyxl.load_workbook(file_path)
        print("The number of worksheets is {0}".format(len(self.book.sheetnames)))
        print("Worksheet name(s): {0}".format(self.book.sheetnames))

        for sheet in self.book.sheetnames:
            self.parent.dlg.refitting_sheet.addItem(sheet)

        self.parent.dlg.refitting_sheet.currentIndexChanged.connect(self.load_xlsx_sheet)


    def load_xlsx_sheet(self):
        """ load data from excel sheet """

        index = self.parent.dlg.refitting_sheet.currentIndex() - 1
        self.sheet = self.book.worksheets[index]
        self.sheet_name = self.book.sheetnames[index]
        self.sheetrows = self.sheet.max_row
        print("Sheet '{0}' has {1} rows and {2} columns".format(self.sheet_name, self.sheet.max_row, self.sheet.max_column))

        # load column indexes
        for column_index in range(self.sheet.max_column):
            cell_obj = self.sheet.cell(1, column_index + 1)
            name = cell_obj.value
            self.col_indexes[name] = column_index
            self.fill_parts(name)
        #print(self.col_indexes)

        self.select_default()

        # load all data into dict in order to save it later to csv file
        self.tablekeys = list(self.col_indexes.keys())
        for rx in range(1, self.sheet.max_row + 1):
            row = {}
            for ry in range(1, self.sheet.max_column + 1):
                col = self.tablekeys[ry-1]
                cell_obj = self.sheet.cell(rx, ry)
                val = cell_obj.value
                type = cell_obj.data_type
                if type == 2: #number
                    val = int(val)
                row[col] = val
            self.table.append(row)
        #print(self.table)


    def load_csv(self, file_path):
        """ load sheets from CSV """

        self.load_csv_sheet(file_path)


    def load_csv_sheet(self, file_path):
        """ load data from CSV sheet """

        with open(file_path, 'r') as csvfile:
            csvreader = csv.reader(csvfile)

            # load column indexes
            self.tablekeys = next(csvreader)
            #print(self.tablekeys)

            i = 0
            for key in self.tablekeys:
                self.fill_parts(key)
                self.col_indexes[key] = i
                i+=1
            #print(self.col_indexes)

            for row in csvreader:
                row_obj = {}
                ry = 0
                for val in row:
                    col = self.tablekeys[ry]
                    if type(val) == int:
                        val = int(val)
                    row_obj[col] = val
                    ry += 1
                self.table.append(row_obj)
            #print(self.table)

            self.sheetrows = csvreader.line_num

        self.sheet_name = ""
        print("Sheet has {0} rows and {1} columns".format(self.sheetrows, len(self.tablekeys)))

        self.select_default()


    def process_refitting(self):
        """ process """

        if self.extension == "csv":
            if not self.utils.check_mandatory_fields(FIELDS_MANDATORY_REMOUNTING_CSV):
                return False
        else:
            if not self.utils.check_mandatory_fields(FIELDS_MANDATORY_REMOUNTING):
                return False

        print("process")

        self.read_parts()
        self.process_parts()
        path = self.write_csv()

        group_name = self.sheet_name
        if group_name == "":
            group_name = os.path.basename(self.parent.dlg.refitting_excel.filePath()).split(".")[0]
        group = self.utils.create_group(group_name)
        self.create_point_layer(path, group)
        self.create_line_layer(path, group)


    def read_parts(self):
        """ read parts row by row """

        for rx in range(1, self.sheetrows):
            #print(rx, self.sheet.row(rx))
            
            part_index = self.col_indexes[self.parent.dlg.refitting_part.currentText()]
            coordx_index = self.col_indexes[self.parent.dlg.refitting_coordx.currentText()]
            coordy_index = self.col_indexes[self.parent.dlg.refitting_coordy.currentText()]
            coordz_index = self.col_indexes[self.parent.dlg.refitting_coordz.currentText()]
            origin_index = self.col_indexes[self.parent.dlg.refitting_origin.currentText()]
            target_index = self.col_indexes[self.parent.dlg.refitting_target.currentText()]
            class_index = self.col_indexes[self.parent.dlg.refitting_class.currentText()]
            labels_index = self.col_indexes[self.parent.dlg.refitting_labels.currentText()]
            colors_index = self.col_indexes[self.parent.dlg.refitting_colors.currentText()]

            if self.extension == "xls":
                part = int(self.sheet.cell_value(rx, part_index))
                coordx = int(self.sheet.cell_value(rx, coordx_index))
                coordy = int(self.sheet.cell_value(rx, coordy_index))
                coordz = int(self.sheet.cell_value(rx, coordz_index))
                origin = int(self.sheet.cell_value(rx, origin_index))
                target = int(self.sheet.cell_value(rx, target_index))
                rclass = int(self.sheet.cell_value(rx, class_index))
                labels = int(self.sheet.cell_value(rx, labels_index))
                color = self.sheet.cell_value(rx, colors_index)
                rx = rx-1

            elif self.extension == "xlsx":
                rx = rx+1
                part = int(self.sheet.cell(rx, part_index+1).value)
                coordx = int(self.sheet.cell(rx, coordx_index+1).value)
                coordy = int(self.sheet.cell(rx, coordy_index+1).value)
                coordz = int(self.sheet.cell(rx, coordz_index+1).value)
                origin = int(self.sheet.cell(rx, origin_index+1).value)
                target = int(self.sheet.cell(rx, target_index+1).value)
                rclass = int(self.sheet.cell(rx, class_index+1).value)
                labels = int(self.sheet.cell(rx, labels_index+1).value)
                color = self.sheet.cell(rx, colors_index+1).value

            elif self.extension == "csv":
                rx -= 1
                cell = self.table[rx]
                part = int(cell[self.parent.dlg.refitting_part.currentText()])
                coordx = int(cell[self.parent.dlg.refitting_coordx.currentText()])
                coordy = int(cell[self.parent.dlg.refitting_coordy.currentText()])
                coordz = int(cell[self.parent.dlg.refitting_coordz.currentText()])
                origin = int(cell[self.parent.dlg.refitting_origin.currentText()])
                target = int(cell[self.parent.dlg.refitting_target.currentText()])
                rclass = int(cell[self.parent.dlg.refitting_class.currentText()])
                labels = int(cell[self.parent.dlg.refitting_labels.currentText()])
                color = cell[self.parent.dlg.refitting_colors.currentText()]

            #print(rx, part, coordx, coordy, coordz, origin, target, rclass, labels)
            self.parts.append({
                "row_num": rx,
                "part_num": part,
                "coordx": coordx,
                "coordy": coordy,
                "coordz": coordz,
                "origin": origin,
                "target": target,
                "rclass": rclass,
                "labels": labels,
                "color": color
            })


    def process_parts(self):
        """ calculate parts """

        for part in self.parts:
            if part["origin"] != 0 and part["target"] != 0:
                self.calculate_refitting(part)


    def calculate_refitting(self, part):
        """ calculate azimut, etc. """

        # inverted origin and target in excel (line before has the right direction)
        origin_num = part["target"]
        target_num = part["origin"]

        origin = self.get_part_by_num(origin_num)
        target = self.get_part_by_num(target_num)

        incx = origin["coordx"] - target["coordx"]
        incy = origin["coordy"] - target["coordy"]
        incxmo = math.sqrt(math.pow(incx, 2))
        incymo = math.sqrt(math.pow(incy, 2))

        origin_point = QgsPointXY(origin["coordx"], origin["coordy"])
        target_point = QgsPointXY(target["coordx"], target["coordy"])

        disthorizontal = QgsDistanceArea().measureLine(origin_point, target_point)
        azimut = self.calculate_azimut(incx, incy, incxmo, incymo)

        eje = azimut
        if azimut > 180:
            eje = azimut - 180

        distvertical = target["coordz"] - origin["coordz"]
        distvertical = abs(distvertical)

        inclinacion = math.atan2(distvertical, disthorizontal)
        inclinacion = 180 * (inclinacion / math.pi)
        inclinacion = abs(inclinacion)

        # print(part["row_num"], part["part_num"], incx, incy, incxmo, incymo, disthorizontal, azimut)

        if self.extension == "xls" or self.extension == "csv":
            row = part["row_num"]
        elif self.extension == "xlsx":
            row = part["row_num"]-1

        self.table[row]["incx"] = incx
        self.table[row]["incy"] = incy
        self.table[row]["incxmo"] = round(incxmo)
        self.table[row]["incymo"] = round(incymo)
        self.table[row]["disthorizontal"] = round(disthorizontal)
        self.table[row]["azimut"] = round(azimut)
        self.table[row]["eje"] = round(eje)
        self.table[row]["distvertical"] = round(distvertical)
        self.table[row]["inclinacion"] = round(inclinacion)

        # save points for later making points layer
        self.save_points_lines(part, origin_point, target_point, origin_num, target_num)


    def calculate_azimut(self, incx, incy, incxmo, incymo):
        """ calculate azimut """

        vatan = 0
        if incx != 0 and incx != 0:
            vatan = math.atan(incx / incy) * (180 / math.pi)

        vazimut = vatan

        if incx >= 0 and incy >= 0:
            vazimut = math.atan(incxmo / incymo) * (180 / math.pi)
        elif incx >= 0 and incy < 0:
            vazimut = 90 + (math.atan(incymo / incxmo) * (180 / math.pi))
        elif incx < 0 and incy < 0:
            vazimut = 180 + (math.atan(incxmo / incymo) * (180 / math.pi))
        elif incx < 0 and incy >= 0:
            vazimut = 270 + (math.atan(incymo / incxmo) * (180 / math.pi))

        return vazimut


    def get_part_by_num(self, num):
        """ calculate azimut """

        for part in self.parts:
            if part["part_num"] == num:
                return part

        return None


    def save_points_lines(self, part, origin_point, target_point, origin_num, target_num):
        """ save points and lines """

        str_pieza_origin = "num"+str(origin_num)
        if not str_pieza_origin in self.points:
            self.points[str_pieza_origin] = {
                "point": origin_point,
                "color": str(part["color"])
            }

        str_pieza_target = "num"+str(target_num)
        if not str_pieza_target in self.points:
            self.points[str_pieza_target] = {
                "point": target_point,
                "color": str(part["color"])
            }

        self.lines.append([origin_point, target_point])
        self.lines_attr.append({
            "origin": int(part["origin"]),
            "target": int(part["target"]),
            "color": str(part["color"]),
            "rclass": int(part["rclass"])
        })


    def clear_parts(self):
        """ empty column combo boxes """

        self.parent.dlg.refitting_part.clear()
        self.parent.dlg.refitting_part.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_coordx.clear()
        self.parent.dlg.refitting_coordx.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_coordy.clear()
        self.parent.dlg.refitting_coordy.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_coordz.clear()
        self.parent.dlg.refitting_coordz.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_origin.clear()
        self.parent.dlg.refitting_origin.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_target.clear()
        self.parent.dlg.refitting_target.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_class.clear()
        self.parent.dlg.refitting_class.addItem(COMBO_SELECT)
        self.parent.dlg.refitting_labels.clear()
        self.parent.dlg.refitting_labels.addItem(COMBO_SELECT)


    def fill_parts(self, name):
        """ fill column combo boxes """

        self.parent.dlg.refitting_part.addItem(name)
        if name.startswith("coord"):
            self.parent.dlg.refitting_coordx.addItem(name)
            self.parent.dlg.refitting_coordy.addItem(name)
            self.parent.dlg.refitting_coordz.addItem(name)
        self.parent.dlg.refitting_origin.addItem(name)
        self.parent.dlg.refitting_target.addItem(name)
        self.parent.dlg.refitting_class.addItem(name)
        self.parent.dlg.refitting_labels.addItem(name)
        self.parent.dlg.refitting_colors.addItem(name)


    def select_default(self):
        """ select default values for column combo boxes """

        for key in FIELDS_DEFAULT:
            widget = self.parent.dlg.tabWidgetMain.widget(2).findChild(QComboBox, key)
            widget.setCurrentText(FIELDS_DEFAULT[key])


    def write_csv(self):
        """ write dict as csv file """

        # save excel sheet as csv
        dir_name = os.path.dirname(self.parent.dlg.refitting_excel.filePath())
        base_path = os.path.basename(self.parent.dlg.refitting_excel.filePath())
        file_path = base_path.split(".")[0]
        file_name = file_path

        if self.sheet_name != "" and self.sheet_name is not None:
            file_name += " - " + self.sheet_name
        path = os.path.join(dir_name, file_name)

        if not os.path.exists(path):
            os.makedirs(path)

        # add info about coords to file name
        coordx_val = self.parent.dlg.refitting_coordx.currentText()[-1]
        coordy_val = self.parent.dlg.refitting_coordy.currentText()[-1]
        if coordx_val and coordy_val:
            file_name += " - " + coordx_val + coordy_val
        file = os.path.join(path, file_name + ".csv")

        with open(file, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.tablekeys)
            if self.extension != "xlsx":
                writer.writeheader()
            writer.writerows(self.table)

            self.parent.dlg.messageBar.pushMessage(f"Refittings done successfully, results saved to file '{file}'", level=Qgis.Success)

            return os.path.dirname(file)


    def create_point_layer(self, path, group):
        """ create line layer """

        layer = self.utils.create_vector_layer("points", "point", group, "&field=id:integer&field=color:string(7)")

        layer.startEditing()
        i=0
        for str_pieza in self.points:
            fields = QgsFields()
            fields.append(QgsField("id", QVariant.Int))
            fields.append(QgsField("color", QVariant.String))
            feature = QgsFeature(fields)
            attr = self.points[str_pieza]
            num_pieza = str_pieza[3:]
            geometry = QgsGeometry.fromPointXY(attr["point"])
            feature.setGeometry(geometry)
            feature.setAttributes([num_pieza, attr["color"]])
            layer.addFeature(feature)
            i+=1
        layer.commitChanges()

        symbology_path = os.path.join(self.parent.plugin_dir, SYMBOLOGY_DIR, "refitting_points.qml")
        layer.loadNamedStyle(symbology_path)
        #layer.triggerRepaint()

        self.utils.save_layer_gpkg(layer, path)


    def create_line_layer(self, path, group):
        """ create point layer """

        layer = self.utils.create_vector_layer("lines", "linestring", group, "&field=origin:integer&field=target:integer&field=color:string(7)&field=class:integer")

        layer.startEditing()
        i=0
        for line in self.lines:
            fields = QgsFields()
            fields.append(QgsField("origin", QVariant.Int))
            fields.append(QgsField("target", QVariant.Int))
            fields.append(QgsField("color", QVariant.String))
            fields.append(QgsField("class", QVariant.Int))
            feature = QgsFeature(fields)
            geometry = QgsGeometry.fromPolylineXY(line)
            feature.setGeometry(geometry)
            attr = self.lines_attr[i]
            feature.setAttributes([attr["origin"], attr["target"], attr["color"], attr["rclass"]])
            layer.addFeature(feature)
            i+=1
        layer.commitChanges()

        symbology_path = os.path.join(self.parent.plugin_dir, SYMBOLOGY_DIR, "refitting_lines.qml")
        layer.loadNamedStyle(symbology_path)
        #layer.triggerRepaint()

        self.utils.save_layer_gpkg(layer, path)
